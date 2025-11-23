"""
Excel Export Service
Сервис для экспорта данных в Excel
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed. Excel export will be disabled.")

from config.settings import settings

logger = logging.getLogger(__name__)


class ExcelService:
    """Сервис для экспорта в Excel"""

    def __init__(self):
        """Инициализация сервиса"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("Excel service initialized without openpyxl")
            return

        self.reports_dir = Path(settings.UPLOAD_DIR) / "exports"
        self.reports_dir.mkdir(parents=True, exist_ok=True)

    def export_requests(
        self,
        requests: List[dict],
        filename: Optional[str] = None
    ) -> str:
        """
        Экспорт запросов в Excel

        Args:
            requests: Список запросов
            filename: Имя файла (опционально)

        Returns:
            str: Путь к созданному файлу
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")

        # Создаем workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Запросы"

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Заголовки
        headers = [
            "ID", "Пользователь", "Тип", "Дата", "Вопрос", "Тип дефекта",
            "Критичность", "Нормативы", "Время обработки", "Из кеша"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Данные
        for row, request in enumerate(requests, 2):
            ws.cell(row=row, column=1).value = request.get('id')
            ws.cell(row=row, column=2).value = f"{request.get('user_first_name', '')} (@{request.get('user_username', 'N/A')})"
            ws.cell(row=row, column=3).value = request.get('request_type', '')
            ws.cell(row=row, column=4).value = request.get('created_at', '')
            ws.cell(row=row, column=5).value = request.get('message_text', '')[:100]  # Первые 100 символов
            ws.cell(row=row, column=6).value = request.get('defect_type', '')
            ws.cell(row=row, column=7).value = request.get('defect_severity', '')
            ws.cell(row=row, column=8).value = ", ".join(request.get('mentioned_regulations', []))
            ws.cell(row=row, column=9).value = f"{request.get('processing_time', 0):.2f}s"
            ws.cell(row=row, column=10).value = "Да" if request.get('cached') else "Нет"

        # Автоширина столбцов
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Сохраняем
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"requests_export_{timestamp}.xlsx"

        filepath = self.reports_dir / filename
        wb.save(str(filepath))

        logger.info(f"Excel export created: {filepath}")
        return str(filepath)

    def export_users(
        self,
        users: List[dict],
        filename: Optional[str] = None
    ) -> str:
        """
        Экспорт пользователей в Excel

        Args:
            users: Список пользователей
            filename: Имя файла

        Returns:
            str: Путь к файлу
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Пользователи"

        # Заголовки
        headers = [
            "ID", "Telegram ID", "Username", "Имя", "Роль",
            "Всего запросов", "Анализов фото", "Голосовых",
            "Дата регистрации", "Последняя активность"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Данные
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1).value = user.get('id')
            ws.cell(row=row, column=2).value = user.get('telegram_id')
            ws.cell(row=row, column=3).value = user.get('username', '')
            ws.cell(row=row, column=4).value = user.get('first_name', '')
            ws.cell(row=row, column=5).value = user.get('role', '')
            ws.cell(row=row, column=6).value = user.get('total_requests', 0)
            ws.cell(row=row, column=7).value = user.get('total_photos', 0)
            ws.cell(row=row, column=8).value = user.get('total_voice', 0)
            ws.cell(row=row, column=9).value = user.get('created_at', '')
            ws.cell(row=row, column=10).value = user.get('last_activity', '')

        # Автоширина
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Сохраняем
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"users_export_{timestamp}.xlsx"

        filepath = self.reports_dir / filename
        wb.save(str(filepath))

        logger.info(f"Users Excel export created: {filepath}")
        return str(filepath)

    def export_analytics(
        self,
        analytics_data: List[dict],
        filename: Optional[str] = None
    ) -> str:
        """
        Экспорт аналитики в Excel

        Args:
            analytics_data: Данные аналитики
            filename: Имя файла

        Returns:
            str: Путь к файлу
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Аналитика"

        # Заголовки
        headers = [
            "Дата", "Период", "Всего запросов", "Всего пользователей",
            "Новых пользователей", "Фото", "Текст", "Голосовые",
            "Дефектов найдено", "Критических", "Значительных", "Незначительных",
            "Среднее время", "Cache hit rate %"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # Данные
        for row, data in enumerate(analytics_data, 2):
            ws.cell(row=row, column=1).value = data.get('date', '')
            ws.cell(row=row, column=2).value = data.get('period_type', '')
            ws.cell(row=row, column=3).value = data.get('total_requests', 0)
            ws.cell(row=row, column=4).value = data.get('total_users', 0)
            ws.cell(row=row, column=5).value = data.get('new_users', 0)
            ws.cell(row=row, column=6).value = data.get('photo_requests', 0)
            ws.cell(row=row, column=7).value = data.get('text_requests', 0)
            ws.cell(row=row, column=8).value = data.get('voice_requests', 0)
            ws.cell(row=row, column=9).value = data.get('defects_found', 0)
            ws.cell(row=row, column=10).value = data.get('critical_defects', 0)
            ws.cell(row=row, column=11).value = data.get('major_defects', 0)
            ws.cell(row=row, column=12).value = data.get('minor_defects', 0)
            ws.cell(row=row, column=13).value = f"{data.get('avg_response_time', 0):.2f}s"
            ws.cell(row=row, column=14).value = f"{data.get('cache_hit_rate', 0):.1f}%"

        # Автоширина
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Сохраняем
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"analytics_export_{timestamp}.xlsx"

        filepath = self.reports_dir / filename
        wb.save(str(filepath))

        logger.info(f"Analytics Excel export created: {filepath}")
        return str(filepath)


# Singleton instance
_excel_service_instance: Optional[ExcelService] = None


def get_excel_service() -> ExcelService:
    """
    Получить экземпляр Excel service (singleton)

    Returns:
        ExcelService: Экземпляр сервиса
    """
    global _excel_service_instance
    if _excel_service_instance is None:
        _excel_service_instance = ExcelService()
    return _excel_service_instance
